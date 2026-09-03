from datetime import date
from unittest import mock

from openpyxl import load_workbook
import io

from reports.export import build_xlsx


def test_build_xlsx_has_two_sheets():
    with mock.patch("reports.export.returns_rows", return_value=[
        {"id": 1, "order_id": 10, "amount": 5.0, "reason": "size", "status": "done", "store": "A"},
    ]), mock.patch("reports.export.cached_returns_by_store", return_value=[
        {"store_id": 1, "cnt": 1, "total": 5.0},
    ]):
        wb = load_workbook(io.BytesIO(build_xlsx(date(2026, 8, 1))))
    assert wb.sheetnames == ["returns", "by_store"]
    assert wb["returns"].max_row == 2
    assert wb["by_store"]["C2"].value == 5.0
